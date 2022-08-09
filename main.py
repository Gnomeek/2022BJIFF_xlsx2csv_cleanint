from dataclasses import dataclass
from datetime import datetime, time
from typing import List

from dataclass_csv import DataclassWriter
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class Movie:
    section: str
    chn_title: str
    en_title: str
    country: str
    year: str
    length: str
    count: int
    price: int
    screen_time: str
    screen_location: str


def spreadsheet_to_cls(wb: Workbook):
    res: List[Movie] = []
    sheet: Worksheet = wb.active
    for row in sheet.iter_rows(min_row=3, values_only=True):
        for column_num in range(9, len(row), 4):
            if not row[column_num]:
                break
            month, day = str(row[column_num]).split(".")
            t: time = row[column_num + 2]
            movie: Movie = Movie(section=row[1], chn_title=row[2], en_title=row[3], country=row[4], year=row[5],
                                 length=row[6], count=int(row[7]), price=int(row[8]),
                                 screen_location=row[column_num + 3],
                                 screen_time=str(datetime(year=2022, month=int(month), day=int(day),
                                                          hour=t.hour, minute=t.minute, second=t.second)))
            res.append(movie)
    return res


if __name__ == '__main__':
    workbook: Workbook = load_workbook(filename="raw.xlsx")
    movies: List[Movie] = spreadsheet_to_cls(workbook)

    with open("cleaned.csv", "w") as f:
        w = DataclassWriter(f, movies, Movie)
        w.write()
